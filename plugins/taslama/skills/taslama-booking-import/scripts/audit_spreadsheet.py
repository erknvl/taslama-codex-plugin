#!/usr/bin/env python3
"""Inventory phone-bearing cells and vertical booking blocks in a calendar-style Excel booking file."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


PHONE_TOKEN = re.compile(r"(?<!\d)(?:\+?993(?:[\s()+-]*\d){8}|8\d{8}|\d{8})(?!\d)")
LETTERS = re.compile(r"[A-Za-zА-Яа-яЁё]")
ADMIN = re.compile(
    r"уборк|дежур|выход|\bвых\b|болеет|больниц|офис|не открывать|напомин|касс",
    re.IGNORECASE,
)


def normalized_phone(token: str) -> str | None:
    digits = "".join(re.findall(r"\d", token))
    if len(digits) == 11 and digits.startswith("993"):
        return "+" + digits
    if len(digits) == 9 and digits.startswith("8"):
        digits = digits[1:]
    if len(digits) == 8:
        return "+993" + digits
    return None


def phones_in(value: object) -> list[str]:
    phones = {
        phone
        for token in PHONE_TOKEN.findall(str(value or ""))
        if (phone := normalized_phone(token))
    }
    return sorted(phones)


def is_descriptive(value: object) -> bool:
    text = str(value or "").strip()
    return bool(text and LETTERS.search(text) and not ADMIN.search(text))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--max-data-row", type=int, default=29)
    parser.add_argument("--max-owner-distance", type=int, default=6)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(args.workbook, data_only=False)
    phone_cells: list[dict[str, object]] = []
    blocks: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        for column in range(2, sheet.max_column + 1):
            professional = str(sheet.cell(2, column).value or "").strip()
            for row in range(3, min(sheet.max_row, args.max_data_row) + 1):
                cell = sheet.cell(row, column)
                phones = phones_in(cell.value)
                if not phones:
                    continue
                owner_cell = cell if is_descriptive(cell.value) else None
                if owner_cell is None:
                    lower_bound = max(3, row - args.max_owner_distance)
                    for owner_row in range(row - 1, lower_bound - 1, -1):
                        candidate = sheet.cell(owner_row, column)
                        if is_descriptive(candidate.value):
                            owner_cell = candidate
                            break
                record = {
                    "sheet": sheet.title,
                    "professional": professional,
                    "column": column,
                    "phoneCell": cell.coordinate,
                    "phoneRaw": str(cell.value),
                    "phones": phones,
                    "ownerCell": owner_cell.coordinate if owner_cell else None,
                    "ownerRaw": str(owner_cell.value) if owner_cell else None,
                    "rowDistance": row - owner_cell.row if owner_cell else None,
                }
                phone_cells.append(record)
                if owner_cell:
                    blocks.append(record)

    summary = {
        "workbook": str(args.workbook),
        "sheets": len(workbook.sheetnames),
        "phoneCells": len(phone_cells),
        "ownedVerticalBlocks": len(blocks),
        "unownedPhoneCells": len(phone_cells) - len(blocks),
    }
    (args.output_dir / "phone-cells.json").write_text(json.dumps(phone_cells, ensure_ascii=False, indent=2), encoding="utf-8")
    (args.output_dir / "vertical-blocks.json").write_text(json.dumps(blocks, ensure_ascii=False, indent=2), encoding="utf-8")
    (args.output_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
